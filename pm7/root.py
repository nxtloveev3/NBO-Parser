from .lib import basicReadingFunctions as brf
from .lib import parseNAOdiscriptor as pNaoD
import copy
import re
import os 
import pandas as pd
import openpyxl
import numpy as np

# These two function extract the information needed and seperate them into different seperate files
# based on the origin file(whether it is a triplet or singlet)
def unrestricted(file, options=None):

    tabs = {}
    
    if 'nboSum' in options:
        start = brf.find("SUMMARY OF NATURAL BOND ORBITALS",file)
        assert(len(start)==2)
        start_alpha, start_beta = start

        end_total = brf.find("TOTAL LEWIS",file)
        end_valence = brf.find("VALENCE NON-NEWIS",file)
        end_alpha, end_beta = 0, 0
        for i in end_total:
            for j in end_valence:
                if j == i+1:
                    if i < start_beta and j < start_beta:
                        end_alpha = max(end_alpha, i)
                    else:
                        assert(i > start_beta and j > start_beta)
                        end_beta = max(end_beta, i)
        tabs['nboSumAlpha'], tabs['nboSumBeta'] = file[start_alpha:end_alpha+5], file[start_beta:end_beta+5]

    if 'nao' in options:
        startNAO = brf.find("NATURAL ATOMIC ORBITAL OCCUPANCIES",file)
        endNAO = brf.find("SUMMARY OF NATURAL POPULATION ANALYSIS",file)
        naoAll = file[startNAO[0]:endNAO[0]]
        naoAlpha = file[startNAO[1]:endNAO[1]]
        naoBeta = file[startNAO[2]:endNAO[2]]
        tabs['naoAll'], tabs['naoAlpha'], tabs['naoBeta'] = naoAll, naoAlpha, naoBeta
    
    return tabs
    
def restricted(file, options=None):

    tabs = {}

    if 'nboSum' in options:
        nboSumStart = brf.find("SUMMARY OF NATURAL BOND ORBITALS (TOTAL DENSITY):",file)
        nboSumEnd = brf.find("NATURAL LOCALIZED MOLECULAR ORBITAL ANALYSIS (TOTAL DENSITY):",file)
        tabs['nboSum'] = file[nboSumStart[0]:nboSumEnd[0]]

    if 'nao' in options:
        startNAO = brf.find("NATURAL ATOMIC ORBITAL OCCUPANCIES (TOTAL DENSITY):",file)
        endNAO = brf.find("SUMMARY OF NATURAL POPULATION ANALYSIS (TOTAL DENSITY):",file)
        tabs['nao'] = file[startNAO[0]:endNAO[0]]
    
    return tabs

class nbo(object):
    def __init__(self, file, option=None, triplet=False, determine_triplet=False):
        lines = brf.readlines(file)
        self.triplet = triplet
        if determine_triplet:
            for line in lines: # Determines if the file is a singlet file or triplet file
                if "alpha spin orbitals" in line:
                    self.triplet = True
                    break
        self.npa,self.badAts,self.badAtsF = findNpa(file)
        if 'npa' in option:
            self.npa = nbo.replacement(self.npa,self.badAts,self.badAtsF)
        else:
            del self.npa
        if self.triplet:
            tabs = unrestricted(lines, options=option)
            if len(re.compile('\W').split(option)) > 1:
                for key in option:
                    if key != 'npa':
                        setattr(self, key+'A', nbo.replacement(tabs[key+'Alpha'],self.badAts,self.badAtsF))
                        setattr(self, key+'B', nbo.replacement(tabs[key+'Beta'],self.badAts,self.badAtsF))
                        if key == 'nao':
                            setattr(self, 'nao', nbo.replacement(tabs['naoAll'],self.badAts,self.badAtsF))
            else:
                if option != 'npa':
                    setattr(self, option, nbo.replacement(tabs[option],self.badAts,self.badAtsF))
                    setattr(self, option, nbo.replacement(tabs[option],self.badAts,self.badAtsF))
                if option == 'nao':
                    setattr(self, 'nao', nbo.replacement(tabs['naoAll'],self.badAts,self.badAtsF))            
        else:
            tabs = restricted(lines, options=option)
            if len(re.compile('\W').split(option)) > 1:
                for key in option:
                    if key != 'npa':
                        setattr(self, key, nbo.replacement(tabs[key],self.badAts,self.badAtsF))
            else:
                if option != 'npa':
                    setattr(self, option, nbo.replacement(tabs[option],self.badAts,self.badAtsF))
    
    #This method replaces all the incorrect characters with correct ones
    @staticmethod
    def replacement(file,incorrect,correct):
        result = file.copy()
        for line in result:
            for num in range(len(incorrect)):
                if incorrect[num] in line:
                    str(result).replace(incorrect[num],correct[num])
        return list(result)


def findNpa(text):
    lines = brf.readlines(text)
    npaStart = brf.find("SUMMARY OF NATURAL POPULATION ANALYSIS (TOTAL DENSITY):",lines)
    pos1 = brf.find("-----------------------------------------------------------------------------",lines)
    pos2 = brf.find("=============================================================================",lines)
    pos11 = []
    pos22 = []
    for elem in pos1:
        if elem > npaStart[0]:
            pos11.append(elem)
            break
    for elem in pos2:
        if elem > pos11[0]:
            pos22.append(elem)
            break
    npa = lines[pos11[0]+1:pos22[0]]
    ats, badAtsF = [line[0] for line in npa], []
    badAts = [atom for atom in ats if not atom.isalpha()]
    for elem in badAts:
        character = ""
        number = ""
        for char in elem:
            if char.isalpha():
                character += char
            else:
                number += char
        result = character + " " + number
        badAtsF.append(result)
    return (npa,badAts,badAtsF)

#This method reparses NPA into a list of dictionaries. Could be converted into a dataframe directly.
def parseNPA(npa, triplet=False) -> list:
    columns = ['Atom', 'No', 'Natural Charge', 'Core', 'Valence', 'Rydeberg', 'Total']
    length = 7
    text = [i.split() for i in npa]
    result = []
    for line in text:
        new = {'Atom': line[1]}
        new['No'] = line[0]
        for i in range(2, length):
            new[columns[i]] = float(line[i])
        result.append(new)
    return result


##### Begin regex constant definition #####
regexInt = r"-?\d+"
regexFloat = r"-?\d+\.\d+"
regexAtom = r"[A-Z][a-z]?"

xyzRegex = brf.namedRe('atomNumber', regexInt, before='allow', after='require')
xyzRegex += brf.namedRe('atom', regexAtom, before='allow', after='require')
xyzRegex += brf.namedRe('x', regexFloat, before='allow', after='require')
xyzRegex += brf.namedRe('y', regexFloat, before='allow', after='require')
xyzRegex += brf.namedRe('z', regexFloat, before='allow', after='allow')
xyxRegexObj = re.compile(xyzRegex)

def parseXYZpm7(filename, verbose=False):
    
    # There are two CARTESIAN COORDINATES table in each .aout file.
    # We only need the *SECOND* geometry xyz table 
    
    lines = brf.readlines(filename)
    xyz_start = brf.find("ATOM            X               Y               Z", lines)[-1]
    lines = lines[xyz_start:]
    lines = lines[:lines.index('')]
    extractedAtoms = [] # list of parsed atom+coord lines

    for line in lines:

        currentLineMatch = xyxRegexObj.fullmatch(line) 
        if currentLineMatch:
            currentLineDict = currentLineMatch.groupdict()
            currentLineDict.pop('atomNumber')
            for key in ['x', 'y', 'z']:
                currentLineDict[key] = float(currentLineDict[key])
            extractedAtoms.append(currentLineDict) # Add matched line to list using a dictionary format.
    
        if verbose:
            print(currentLineMatch, ' ', line)

    return extractedAtoms

def discriptorParserNpa(fileS, fileT, verbose = False):
    fdS = openpyxl.load_workbook(fileS)
    fdT = openpyxl.load_workbook(fileT)
    sheetS = fdS['Sheet1']
    sheetT = fdT['Sheet1']
    #if type == None: print("Please provide type singlet/triplet")
    npa = ['Natural charge Singlet', 'Natural charge Triplet']  #'Natural Spin Density'
    for i in range(len(npa)):
        npa[i] = []

    for row in range(2, sheetS.max_row + 1):
        chargeS = sheetS['D' + str(row)].value
        npa[0].append(chargeS)
    for row in range(2, sheetT.max_row + 1):
        chargeT = sheetT['D' + str(row)].value    
        npa[1].append(chargeT)
    #if type == 'singlet':
        #result = pd.DataFrame(npa)
        #result.fillna(0)
        #if verbose: print(result)
        #return result
    #elif type == 'triplet' and text != None:
        #lines = brf.readlines(text)
        # pos1 = brf.find("ATOMIC ORBITAL SPIN POPULATIONS", lines)
        #pos2 = brf.find("N A T U R A L   A T O M I C   O R B I T A L   A N D", lines)
        #spinDens = lines[(pos1[0]+1):(pos2[0]-1)]
        #for num in spinDens:
            #for float in re.split(' +', num):
                #npa[2].append(float)
    #else: print("Require text for npa spin population!")   
    result = pd.DataFrame(npa)
    result.fillna(0) 
    if verbose: print(result)
    return result.T

def parseDiscriptor(pathS, pathT):
    disDic = {}
    for filename in os.listdir(pathS):
        if filename.endswith(".xlsx"):
            complex = re.findall(r'\d+', filename)
            assert(len(complex) == 2)
            cn = complex[0]
            nn = complex[1]
            if cn not in disDic:
                disDic[cn] = {}
            disDic[cn][nn] = {}
            try:
                disDic[cn][nn]['npa'] = discriptorParserNpa(os.path.join(pathS, filename),os.path.join(pathT, filename), verbose = False)
            except:
                print(filename)
    #for filename in os.listdir(pathT):
        #if filename.endswith(".xlsx"):
            #complex = re.findall(r'\d+', filename)
            #assert(len(complex) == 2)
            #cn = complex[0]
            #nn = complex[1]
            #for text in os.listdir(filePath):
                #if cn in text and nn in text:
                    #textName = text
            #assert(textName != None)
            #if cn not in disDic or nn not in disDic[cn]:
                #print(cn, nn, 'not found in singlet file')
            #else:
                #singlet = disDic[cn][nn]['npa']
                #try:
                    #triplet = discriptorParserNpa(os.path.join(pathT, filename), type = 'triplet', verbose = False, text = os.path.join(filePath, textName))
                    #npa = singlet.append(triplet)
                    #disDic[cn][nn]['npa'] = npa.T
                #except:
                    #print(filename, textName)
    #disDic[cn][nn]['nao'] = pNaoD.discriptorParserNao(os.path.join(pathS, filename), option = 'naoS', verbose = False)
    #disDic[cn][nn]['nao'] = pNaoD.discriptorParserNao(os.path.join(path, filename), option = 'naoT', verbose = False)
    return disDic
                